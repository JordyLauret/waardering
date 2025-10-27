#!/usr/bin/env python3
"""
Waarderingsrapport-generator (MVP)
- Leest Excel met aannames en berekeningen
- Maakt een klantvriendelijk PDF-rapport (DCF + multiples placeholders)
"""
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from reportlab.lib.units import cm
import datetime
import textwrap

def wrap(text, w=110):
    import textwrap
    return "<br/>".join(textwrap.wrap(text, width=w))

def df_to_table(df, col_widths=None):
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    data = [list(df.columns)] + df.astype(object).where(pd.notnull(df), "").values.tolist()
    if col_widths is None:
        t = Table(data, repeatRows=1)
    else:
        t = Table(data, repeatRows=1, colWidths=col_widths)
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0), colors.lightgrey),
        ("GRID",(0,0),(-1,-1), 0.25, colors.grey),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("ALIGN",(1,1),(-1,-1),"RIGHT")
    ]))
    return t

def analyze_workbook(path: Path):
    xls = pd.ExcelFile(path)
    sheet_names = xls.sheet_names
    # profiles
    profiles = []
    samples = {}
    for name in sheet_names:
        try:
            df = pd.read_excel(xls, sheet_name=name, header=0)
            rows, cols = df.shape
            non_null = int(df.notna().sum().sum())
            profiles.append({"Tabblad": name, "Rijen": rows, "Kolommen": cols, "Niet-lege cellen": non_null})
            samples[name] = df.head(5)
        except Exception:
            profiles.append({"Tabblad": name, "Rijen": None, "Kolommen": None, "Niet-lege cellen": None})
    profiles_df = pd.DataFrame(profiles)

    # formulas and names
    wb = load_workbook(path, data_only=False, read_only=True)
    defined_names = [dn.name for dn in wb.defined_names.definedName]

    formula_overview = []
    for wsname in sheet_names:
        ws = wb[wsname]
        formula_count = 0
        total_cells = 0
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if cell is None:
                    continue
                total_cells += 1
                val = cell.value
                if isinstance(val, str) and str(val).startswith("="):
                    formula_count += 1
        formula_overview.append({"Tabblad": wsname, "Formulecellen": formula_count, "Gesamplede cellen": total_cells})
    formula_df = pd.DataFrame(formula_overview).sort_values("Formulecellen", ascending=False)

    # params candidates
    param_candidates = []
    param_data = {}
    xls2 = pd.ExcelFile(path)
    for name in sheet_names:
        try:
            df = pd.read_excel(xls2, sheet_name=name, header=0)
            if 2 <= df.shape[1] <= 6 and len(df) <= 500:
                types = df.dtypes.astype(str).tolist()
                if any("object" in t for t in types) and any(("int" in t or "float" in t) for t in types):
                    non_empty = int(df.iloc[:,0].notna().sum())
                    param_candidates.append({"Tabblad": name, "Rijen": int(len(df)), "Kolommen": int(df.shape[1]), "Niet-lege eerste kolom": non_empty})
                    param_data[name] = df.head(10)
        except Exception:
            pass
    param_df = pd.DataFrame(param_candidates).sort_values(["Rijen","Kolommen"], ascending=[True, True])

    # detect likely series
    keywords = {
        "revenue": ["revenue","omzet","sales","turnover"],
        "ebitda": ["ebitda"],
        "ebit": ["ebit","bedrijfsresultaat"],
        "netincome": ["net income","nettowinst","resultaat na belasting","winst"],
        "capex": ["capex","invest","investeringen"],
        "wc": ["werkkapitaal","working capital","net working capital","nwc"],
    }
    detected = {}
    for sheet, df in samples.items():
        cols = [str(c).strip().lower() for c in df.columns]
        for k, keys in keywords.items():
            for key in keys:
                for c in cols:
                    if key in c:
                        detected.setdefault(sheet, set()).add((k, c))
    detected = {k:list(v) for k,v in detected.items()}
    return profiles_df, formula_df, param_df, defined_names, detected

def generate_pdf(input_xlsx: Path, output_pdf: Path, client_name="Klantnaam", valuation_date=None):
    profiles_df, formula_df, param_df, defined_names, detected_series = analyze_workbook(input_xlsx)
    if valuation_date is None:
        import datetime
        valuation_date = datetime.date.today()

    doc = SimpleDocTemplate(str(output_pdf), pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Header", fontSize=16, leading=20, spaceAfter=12))
    styles.add(ParagraphStyle(name="Subheader", fontSize=12, leading=16, spaceAfter=8))
    styles.add(ParagraphStyle(name="Small", fontSize=9, leading=12))
    styles.add(ParagraphStyle(name="Emph", fontSize=10, leading=14, spaceAfter=6, textColor=colors.black))
    content = []

    title = f"Waardering – {client_name}"
    subtitle = f"Rapportdatum: {valuation_date.strftime('%d-%m-%Y')}"
    content.append(Paragraph(title, styles["Title"]))
    content.append(Spacer(1, 6))
    content.append(Paragraph(subtitle, styles["Normal"]))
    content.append(Spacer(1, 18))
    content.append(Paragraph("Samenvatting (Management)", styles["Header"]))
    content.append(Paragraph(wrap("Executive summary met bandbreedte, kern­aannames en duiding (placeholder in MVP)."), styles["Normal"]))

    content.append(Spacer(1, 12))
    content.append(Paragraph("1. Methodiek", styles["Header"]))
    content.append(Paragraph(wrap("DCF + multiples; kasstromen uit Excel, discontering met WACC; correcties naar equity value."), styles["Normal"]))

    content.append(Spacer(1, 6))
    content.append(Paragraph("1.1 Aannames & invoer (uit Excel)", styles["Subheader"]))
    if not param_df.empty:
        content.append(df_to_table(param_df, [7*cm, 2.5*cm, 2.5*cm, 3*cm]))
    else:
        content.append(Paragraph("Geen duidelijke parameter-tabbladen gevonden (heuristisch).", styles["Small"]))

    content.append(Spacer(1, 12))
    content.append(Paragraph("2. Projecties en kasstromen", styles["Header"]))
    if detected_series:
        for sheet, pairs in detected_series.items():
            content.append(Paragraph(f"Herkenning in tabblad: <b>{sheet}</b>", styles["Subheader"]))
            for k, c in pairs:
                content.append(Paragraph(f"• {k} → kolom '{c}'", styles["Small"]))
    else:
        content.append(Paragraph("Geen relevante kolomnamen automatisch gedetecteerd.", styles["Small"]))

    content.append(Spacer(1, 12))
    content.append(Paragraph("3. DCF-berekening", styles["Header"]))
    content.append(Paragraph("Formules en resultaten volgen in de definitieve versie (placeholder).", styles["Small"]))

    content.append(Spacer(1, 12))
    content.append(Paragraph("4. Multiples-benadering", styles["Header"]))
    content.append(Paragraph("Peers en toegepaste multiple volgen in de definitieve versie (placeholder).", styles["Small"]))

    content.append(Spacer(1, 12))
    content.append(Paragraph("5. Scenario’s en gevoeligheid", styles["Header"]))
    content.append(Paragraph("Scenario- en gevoeligheidstabellen verschijnen hier (placeholder).", styles["Small"]))

    content.append(PageBreak())
    content.append(Paragraph("Bijlage A – Werkmap-overzicht", styles["Header"]))
    if not profiles_df.empty:
        content.append(df_to_table(profiles_df, [6*cm, 2.5*cm, 2.5*cm, 3*cm]))

    content.append(Spacer(1, 10))
    content.append(Paragraph("Bijlage B – Formule-overzicht per tabblad", styles["Header"]))
    if not formula_df.empty:
        content.append(df_to_table(formula_df, [8*cm, 3.5*cm, 3.5*cm]))

    content.append(Spacer(1, 10))
    content.append(Paragraph("Bijlage C – Gedefinieerde namen (Excel)", styles["Header"]))
    if defined_names:
        batch = ", ".join(sorted(defined_names)[:120])
        content.append(Paragraph(wrap(batch, 110), styles["Small"]))
    else:
        content.append(Paragraph("Geen gedefinieerde namen aangetroffen.", styles["Normal"]))

    doc.build(content)

if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--input", required=True, help="Pad naar Excel")
    p.add_argument("--output", required=True, help="Pad naar uitgaand PDF")
    p.add_argument("--client", default="Klantnaam", help="Klantnaam in titel")
    args = p.parse_args()
    generate_pdf(Path(args.input), Path(args.output), client_name=args.client)
